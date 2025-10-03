# src/reports/excel_spc_export_generator.py - COMPLETE WITH ALL IMPROVEMENTS
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, Any
import logging

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Border, Side, Alignment, NamedStyle)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins

from src.models.plotting.spc_data_loader import SPCDataLoader
from src.gui.logging_config import logger as base_logger


class ExcelSPCReportGenerator:
    """Professional Excel report generator with complete data"""
    
    COLORS = {
        'primary_blue': '1B365D',
        'secondary_blue': '2E5266',
        'accent_blue': '4A90A4',
        'white': 'FFFFFF',
        'light_gray': 'F7F8F9',
        'medium_gray': 'E1E5E9',
        'dark_gray': '2C3E50',
        'success_green': '27AE60',
        'warning_orange': 'F39C12',
        'danger_red': 'E74C3C',
        'table_header': 'ECF0F1',
        'table_alt': 'F8F9FA',
        'excellent_teal': '1ABC9C'
    }
    
    CHART_TYPES = {
        'normality': 'NORMALITY ANALYSIS',
        'capability': 'PROCESS CAPABILITY',
        'individuals': 'I-CHART (INDIVIDUALS)',
        'moving_range': 'MR-CHART (MOVING RANGE)',
        'xbar': 'X̄-CHART (AVERAGE)',
        'r_chart': 'R-CHART (RANGE)',
        's_chart': 'S-CHART (STD DEV)',
        'distribution': 'DISTRIBUTION ANALYSIS'
    }

    def __init__(
        self,
        client: str,
        ref_project: str,
        batch_number: str,
        base_path: str = "./data/spc",
        charts_path: str = "./data/reports/charts",
        output_path: str = "./data/reports/excel",
        logger: Optional[logging.Logger] = None
    ):
        self.client = client
        self.ref_project = ref_project
        self.batch_number = batch_number
        self.base_path = Path(base_path)
        self.charts_path = Path(charts_path)
        self.output_path = Path(output_path)
        self.logger = logger or base_logger.getChild(self.__class__.__name__)
        
        self.output_path.mkdir(parents=True, exist_ok=True)
        
        self.elements_data = {}
        self.workbook = None
        self.styles_created = False
        
        self.logger.info(f"Initialized Excel SPC Report Generator for {client}_{ref_project}_{batch_number}")

    def create_professional_styles(self):
        """Create professional styles"""
        if self.styles_created:
            return
            
        doc_title = NamedStyle(name="doc_title")
        doc_title.font = Font(name='Arial', size=20, bold=True, color=self.COLORS['primary_blue'])
        doc_title.alignment = Alignment(horizontal='center', vertical='center')
        
        doc_subtitle = NamedStyle(name="doc_subtitle")
        doc_subtitle.font = Font(name='Arial', size=12, color=self.COLORS['dark_gray'])
        doc_subtitle.alignment = Alignment(horizontal='center', vertical='center')
        
        main_header = NamedStyle(name="main_header")
        main_header.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        main_header.fill = PatternFill(start_color=self.COLORS['primary_blue'], 
                                     end_color=self.COLORS['primary_blue'], 
                                     fill_type='solid')
        main_header.alignment = Alignment(horizontal='center', vertical='center')
        main_header.border = self.create_full_border('medium', self.COLORS['primary_blue'])
        
        section_header = NamedStyle(name="section_header")
        section_header.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        section_header.fill = PatternFill(start_color=self.COLORS['secondary_blue'], 
                                        end_color=self.COLORS['secondary_blue'], 
                                        fill_type='solid')
        section_header.alignment = Alignment(horizontal='center', vertical='center')
        section_header.border = self.create_full_border('thin', self.COLORS['medium_gray'])
        
        table_header = NamedStyle(name="table_header")
        table_header.font = Font(name='Arial', size=9, bold=True, color=self.COLORS['dark_gray'])
        table_header.fill = PatternFill(start_color=self.COLORS['table_header'], 
                                      end_color=self.COLORS['table_header'], 
                                      fill_type='solid')
        table_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        table_header.border = self.create_full_border('thin', self.COLORS['medium_gray'])
        
        param_label = NamedStyle(name="param_label")
        param_label.font = Font(name='Arial', size=9, bold=True, color=self.COLORS['dark_gray'])
        param_label.fill = PatternFill(start_color=self.COLORS['light_gray'], 
                                     end_color=self.COLORS['light_gray'], 
                                     fill_type='solid')
        param_label.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        param_label.border = self.create_full_border('thin', self.COLORS['medium_gray'])
        
        data_cell = NamedStyle(name="data_cell")
        data_cell.font = Font(name='Arial', size=9, color=self.COLORS['dark_gray'])
        data_cell.alignment = Alignment(horizontal='center', vertical='center')
        data_cell.border = self.create_full_border('thin', self.COLORS['medium_gray'])
        
        chart_title = NamedStyle(name="chart_title")
        chart_title.font = Font(name='Arial', size=11, bold=True, color=self.COLORS['secondary_blue'])
        chart_title.alignment = Alignment(horizontal='center', vertical='center')
        chart_title.fill = PatternFill(start_color=self.COLORS['light_gray'], 
                                     end_color=self.COLORS['light_gray'], 
                                     fill_type='solid')
        chart_title.border = self.create_full_border('thin', self.COLORS['medium_gray'])
        
        notes_style = NamedStyle(name="notes_style")
        notes_style.font = Font(name='Arial', size=9, color=self.COLORS['dark_gray'])
        notes_style.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        notes_style.border = self.create_full_border('thin', self.COLORS['medium_gray'])
        
        try:
            styles_to_add = [
                doc_title, doc_subtitle, main_header, section_header,
                table_header, param_label, data_cell, chart_title, notes_style
            ]
            
            for style in styles_to_add:
                self.workbook.add_named_style(style)
            
            self.styles_created = True
            self.logger.debug("Created all professional Excel styles")
            
        except Exception as e:
            self.logger.error(f"Error creating styles: {e}")

    def create_full_border(self, style: str = 'thin', color: str = None) -> Border:
        """Create a complete border for all sides"""
        if color is None:
            color = self.COLORS['medium_gray']
        
        return Border(
            left=Side(style=style, color=color),
            right=Side(style=style, color=color),
            top=Side(style=style, color=color),
            bottom=Side(style=style, color=color)
        )

    def apply_complete_borders_to_range(self, ws: Worksheet, start_cell: str, end_cell: str):
        """Apply complete borders to a range of cells"""
        for row in ws[f"{start_cell}:{end_cell}"]:
            for cell in row:
                cell.border = self.create_full_border('thin', self.COLORS['medium_gray'])

    def extract_element_name(self, element_key: str) -> str:
        """Extract clean element name without cavity information"""
        if " Cavity " in element_key:
            return element_key.split(" Cavity ")[0]
        elif "_cavity_" in element_key:
            return element_key.split("_cavity_")[0]
        elif " cavity " in element_key.lower():
            return element_key.lower().split(" cavity ")[0]
        return element_key

    def load_data(self) -> bool:
        """Load SPC data for the report"""
        try:
            folder_name = f"{self.client}_{self.ref_project}_{self.batch_number}"
            filename = f"{self.ref_project}_{self.batch_number}_complete_report.json"
            report_path = self.base_path / folder_name / filename
            
            if not report_path.exists():
                alt_folder_name = f"{self.client}_{self.ref_project}"
                alt_report_path = self.base_path / alt_folder_name / filename
                
                if alt_report_path.exists():
                    report_path = alt_report_path
                else:
                    self.logger.error("SPC report not found")
                    return False
            
            data_loader = SPCDataLoader(self.ref_project, self.base_path)
            self.elements_data = data_loader.load_complete_report(report_path)
            
            if not self.elements_data:
                self.logger.error("No elements data loaded")
                return False
                
            self.logger.info(f"Loaded data for {len(self.elements_data)} elements")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading data: {e}")
            return False

    def get_chart_path(self, element_key: str, chart_type: str) -> Optional[Path]:
        """Get the path to a specific chart file"""
        try:
            element_data = self.elements_data.get(element_key, {})
            original_element_name = element_data.get("element_name")
            cavity = element_data.get("cavity", "")
            
            if not original_element_name:
                original_element_name = self.extract_element_name(element_key)
            
            if cavity and str(cavity).strip():
                filename = f"{chart_type}_{self.batch_number}_{original_element_name}_{cavity}.png"
            else:
                filename = f"{chart_type}_{self.batch_number}_{original_element_name}.png"
            
            chart_path = self.charts_path / filename
            
            if chart_path.exists():
                return chart_path
            else:
                self.logger.warning(f"Chart not found: {chart_path}")
                return None
                
        except Exception as e:
            self.logger.error(f"Error getting chart path: {e}")
            return None

    def create_report(self, 
                    part_description: str = "",
                    drawing_number: str = "",
                    methodology: str = "CMM",
                    facility: str = "",
                    dimension_class: str = "CC") -> str:
        """Create the complete professional Excel SPC report"""
        try:
            if not self.load_data():
                raise ValueError("Failed to load SPC data")
            
            self.workbook = Workbook()
            
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            
            self.create_professional_styles()
            
            self.create_executive_summary_sheet()
            
            sheets_created = 0
            for element_key, element_data in self.elements_data.items():
                try:
                    sheet_name = element_key[:31] if len(element_key) > 31 else element_key
                    invalid_chars = ['/', '\\', '[', ']', '*', '?', ':']
                    for char in invalid_chars:
                        sheet_name = sheet_name.replace(char, '_')
                    
                    ws = self.workbook.create_sheet(title=sheet_name)
                    
                    self.create_professional_element_report(
                        ws, element_key, element_data,
                        part_description, drawing_number, 
                        methodology, facility, dimension_class
                    )
                    
                    sheets_created += 1
                    self.logger.info(f"✓ Created sheet for: {element_key}")
                    
                except Exception as e:
                    self.logger.error(f"✗ Failed sheet for {element_key}: {e}")
                    continue
            
            if sheets_created == 0:
                raise ValueError("No sheets were created successfully")
            
            filename = f"{self.client}_{self.ref_project}_{self.batch_number}_SPC_Analysis_Report.xlsx"
            output_file = self.output_path / filename
            
            self.workbook.save(output_file)
            self.logger.info(f"✅ Report created: {output_file}")
            self.logger.info(f"📊 Contains: 1 summary + {sheets_created} element sheets")
            
            return str(output_file)
            
        except Exception as e:
            self.logger.error(f"❌ Error creating Excel report: {e}")
            raise

    def create_professional_element_report(self, 
                                         ws: Worksheet, 
                                         element_key: str, 
                                         element_data: Dict[str, Any],
                                         part_description: str,
                                         drawing_number: str,
                                         methodology: str,
                                         facility: str,
                                         dimension_class: str):
        """Create professional report for one element with ALL VALUES"""
        
        current_row = 1
        
        # Header with metadata
        current_row = self.create_professional_header(
            ws, current_row, element_key, part_description,
            drawing_number, methodology, facility, dimension_class, element_data
        )       
        
        # First page charts
        current_row = self.create_first_page_charts_optimized(ws, current_row, element_key)
        
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1
        
        # Second page with control charts and values
        current_row = self.create_second_page_with_values(ws, current_row, element_key, element_data)
        
        self.setup_professional_page_formatting(ws)

    def create_professional_header(self, ws, start_row, element_key, part_description,
                                drawing_number, methodology, facility, dimension_class, element_data):
        """Create header WITH all metadata properly transmitted"""
        
        current_row = start_row
        
        # Logo
        logo_path = Path("./assets/images/gui/logo_some.png")
        if logo_path.exists():
            try:
                img = Image(str(logo_path))
                img.width = 245
                img.height = 57
                ws.add_image(img, f'A{current_row}')
            except Exception as e:
                self.logger.warning(f"Could not add logo: {e}")
        
        # Title
        ws.merge_cells(f'C{current_row}:H{current_row}')
        title_cell = ws[f'C{current_row}']
        title_cell.value = "STATISTICAL PROCESS CONTROL"
        title_cell.style = 'doc_title'
        ws.row_dimensions[current_row].height = 33
        
        current_row += 1
        ws.merge_cells(f'C{current_row}:H{current_row}')
        subtitle_cell = ws[f'C{current_row}']
        subtitle_cell.value = "Capability Analysis Report"
        subtitle_cell.style = 'doc_subtitle'
        ws.row_dimensions[current_row].height = 25
        
        current_row += 1
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        
        # Header row
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = "PART & PROJECT INFORMATION"
        header_cell.style = 'main_header'
        ws.row_dimensions[current_row].height = 30
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        
        current_row += 1
        
        # Get element data - PROPERLY EXTRACT CLASS, INSTRUMENT, SIGMA
        clean_element_name = self.extract_element_name(element_key)
        cavity_info = element_data.get('cavity', 'N/A')
        
        # CRITICAL: Extract from element_data (these come from the study)
        element_class = element_data.get('class', dimension_class)
        instrument = element_data.get('instrument', 'CMM')
        sigma_level = element_data.get('sigma', '6σ')
        
        # Determine PP/CP requirement based on ACTUAL sigma
        sigma_numeric = 6 if '6' in str(sigma_level) else 5
        required_pp = 2.0 if sigma_numeric == 6 else 1.67
        pp = element_data.get('pp', 0)
        meets_requirement = "YES ✓" if pp >= required_pp else "NO ✗"
        
        info_data = [
            ["Client:", self.client, "Part Description:", part_description],
            ["Project Reference:", self.ref_project, "Drawing Number:", drawing_number],
            ["Batch Number:", self.batch_number, "Methodology:", methodology],
            ["Quality Facility:", facility, "Dimension Class:", element_class],
            ["Element Name:", clean_element_name, "Cavity:", cavity_info],
            ["Measuring Instrument:", instrument, "Sigma Level:", sigma_level],
            ["PP Requirement:", f"≥{required_pp:.2f}", "Meets Req.:", meets_requirement],
        ]
        
        for row_data in info_data:
            ws.row_dimensions[current_row].height = 25

            ws[f'A{current_row}'].value = row_data[0]
            ws[f'A{current_row}'].style = 'param_label'
            ws[f'E{current_row}'].value = row_data[2]
            ws[f'E{current_row}'].style = 'param_label'
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'].value = row_data[1]
            ws[f'B{current_row}'].style = 'data_cell'
            
            ws.merge_cells(f'F{current_row}:H{current_row}')
            ws[f'F{current_row}'].value = row_data[3]
            ws[f'F{current_row}'].style = 'data_cell'
            
            # Color code meets requirement
            if "Meets Req." in row_data[2]:
                cell = ws[f'F{current_row}']
                if "YES" in row_data[3]:
                    cell.fill = PatternFill(start_color=self.COLORS['success_green'], 
                                        end_color=self.COLORS['success_green'], fill_type='solid')
                    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                else:
                    cell.fill = PatternFill(start_color=self.COLORS['danger_red'], 
                                        end_color=self.COLORS['danger_red'], fill_type='solid')
                    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            
            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
            current_row += 1
        
        ws.row_dimensions[current_row].height = 15
        current_row += 1
        return current_row

    def create_statistical_summary_table(self, 
                                        ws: Worksheet, 
                                        start_row: int, 
                                        element_data: Dict[str, Any]) -> int:
        """Create statistical summary WITH AD statistic and p-value for ORIGINAL sample"""
        
        current_row = start_row
        
        # Section header
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "STATISTICAL ANALYSIS RESULTS"
        section_cell.style = 'section_header'
        ws.row_dimensions[current_row].height = 28
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        
        current_row += 1
        
        # Table headers
        ws.row_dimensions[current_row].height = 24
        
        ws[f'A{current_row}'].value = "Parameter"
        ws[f'A{current_row}'].style = 'table_header'
        ws[f'E{current_row}'].value = "Parameter"
        ws[f'E{current_row}'].style = 'table_header'
        
        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws[f'B{current_row}'].value = "Value"
        ws[f'B{current_row}'].style = 'table_header'
        
        ws.merge_cells(f'F{current_row}:H{current_row}')
        ws[f'F{current_row}'].value = "Value"
        ws[f'F{current_row}'].style = 'table_header'
        
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        current_row += 1
        
        # Parameters - INCLUDING AD and p-value
        left_params = [
            ('sample_size', 'Sample Size (n)', ''),
            ('nominal', 'Nominal Target', '.3f'),
            ('tolerance', 'Tolerance Range', '.3f'),
            ('mean', 'Mean (X̄)', '.4f'),
            ('std_long', 'Long-term Std Dev (σ)', '.4f'),
            ('std_short', 'Short-term Std Dev (s)', '.4f'),
            ('ad_value', 'Anderson-Darling (A²)', '.3f'),  # ADDED
        ]
        
        right_params = [
            ('cp', 'CP', '.3f'),
            ('cpk', 'CPK', '.3f'),
            ('pp', 'PP', '.3f'),
            ('ppk', 'PPK', '.3f'),
            ('ppm_short', 'PPM Defective (Short-term)', '.0f'),
            ('ppm_long', 'PPM Defective (Long-term)', '.0f'),
            ('p_value', 'Normality p-value', '.4f'),  # ADDED
        ]
        
        max_rows = max(len(left_params), len(right_params))
        
        for i in range(max_rows):
            ws.row_dimensions[current_row].height = 20

            # Left side
            if i < len(left_params):
                key, label, fmt = left_params[i]
                if key:
                    ws[f'A{current_row}'].value = label
                    ws[f'A{current_row}'].style = 'param_label'
                    
                    value = element_data.get(key, 'N/A')
                    if isinstance(value, (int, float)) and value != 'N/A' and fmt:
                        formatted_value = f"{value:{fmt}}"
                    else:
                        formatted_value = str(value)
                    
                    ws.merge_cells(f'B{current_row}:D{current_row}')
                    ws[f'B{current_row}'].value = formatted_value
                    ws[f'B{current_row}'].style = 'data_cell'
            
            # Right side  
            if i < len(right_params):
                key, label, fmt = right_params[i]
                if key:
                    ws[f'E{current_row}'].value = label
                    ws[f'E{current_row}'].style = 'param_label'
                    
                    value = element_data.get(key, 'N/A')
                    if isinstance(value, (int, float)) and value != 'N/A' and fmt:
                        formatted_value = f"{value:{fmt}}"
                    else:
                        formatted_value = str(value)
                    
                    ws.merge_cells(f'F{current_row}:H{current_row}')
                    ws[f'F{current_row}'].value = formatted_value
                    ws[f'F{current_row}'].style = 'data_cell'
            
            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
            current_row += 1
        
        ws.row_dimensions[current_row].height = 10
        current_row += 1
        return current_row

    def create_first_page_charts_optimized(self, ws: Worksheet, start_row: int, element_key: str) -> int:
        """Create first page charts"""
        
        current_row = start_row
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        charts_header = ws[f'A{current_row}']
        charts_header.value = "STATISTICAL CONTROL CHARTS"
        charts_header.style = 'section_header'
        ws.row_dimensions[current_row].height = 25
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        current_row += 1
        
        # Normality
        normality_path = self.get_chart_path(element_key, 'normality')
        if normality_path:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'].value = self.CHART_TYPES['normality']
            ws[f'A{current_row}'].style = 'chart_title'
            ws.row_dimensions[current_row].height = 25
            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
            current_row += 1
            
            ws.row_dimensions[current_row].height = 8
            current_row += 1
            
            current_row = self.add_centered_chart_optimized(ws, normality_path, current_row)
            
            ws.row_dimensions[current_row].height = 15
            current_row += 1
        
        # Capability
        capability_path = self.get_chart_path(element_key, 'capability')
        if capability_path:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'].value = self.CHART_TYPES['capability']
            ws[f'A{current_row}'].style = 'chart_title'
            ws.row_dimensions[current_row].height = 25
            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
            current_row += 1

            ws.row_dimensions[current_row].height = 7  # CHANGED from default to 7.5
            #current_row += 1
            
            current_row = self.add_centered_chart_optimized(ws, capability_path, current_row)
            current_row += 1
        
        return current_row

    def create_second_page_with_values(self, ws, start_row, element_key, element_data):
        """Create second page with control charts, distribution, AND ALL MEASURED VALUES"""
        
        # ADJUSTED: Changed from "start_row + 1" to just "start_row"
        # And set the first spacing row to 13 pixels (9.75 points)
        current_row = start_row

        ws.row_dimensions[current_row].height = 10  # 10 pixels - the adjusted spacing row
        
        # Control charts
        xbar_path = self.get_chart_path(element_key, 'xbar')
        r_path = self.get_chart_path(element_key, 'r_chart')
        s_path = self.get_chart_path(element_key, 's_chart')
        individuals_path = self.get_chart_path(element_key, 'individuals')
        mr_path = self.get_chart_path(element_key, 'moving_range')
        
        if xbar_path and r_path:
            left_chart = xbar_path
            right_chart = r_path
            left_title = "X̄-CHART (AVERAGE)"
            right_title = "R-CHART (RANGE)"
        elif xbar_path and s_path:
            left_chart = xbar_path
            right_chart = s_path
            left_title = "X̄-CHART (AVERAGE)"
            right_title = "S-CHART (STD DEV)"
        elif individuals_path and mr_path:
            left_chart = individuals_path
            right_chart = mr_path
            left_title = "I-CHART (INDIVIDUALS)"
            right_title = "MR-CHART (MOVING RANGE)"
        else:
            left_chart = right_chart = None
        
        if left_chart and right_chart:
            ws.row_dimensions[current_row].height = 25
            ws.merge_cells(f'A{current_row}:D{current_row}')
            ws[f'A{current_row}'].value = left_title
            ws[f'A{current_row}'].style = 'chart_title'
            
            ws.merge_cells(f'E{current_row}:H{current_row}')
            ws[f'E{current_row}'].value = right_title
            ws[f'E{current_row}'].style = 'chart_title'
            current_row += 1
            
            ws.row_dimensions[current_row].height = 8
            current_row += 1
            
            chart_row = current_row
            self.add_side_by_side_chart_optimized(ws, left_chart, chart_row, 'left')
            self.add_side_by_side_chart_optimized(ws, right_chart, chart_row, 'right')
            
            current_row = chart_row + 13
            current_row += 1

        # Distribution chart
        distribution_path = self.get_chart_path(element_key, 'distribution')
        if distribution_path:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'].value = "DISTRIBUTION ANALYSIS"
            ws[f'A{current_row}'].style = 'chart_title'
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            
            ws.row_dimensions[current_row].height = 8
            current_row += 1
            
            current_row = self.add_centered_chart_optimized(ws, distribution_path, current_row)
            ws.row_dimensions[current_row].height = 15
            current_row += 1
        
        # CRITICAL: Add ALL MEASURED VALUES section
        current_row = self.create_measured_values_section(ws, current_row, element_data)
        
        # Statistical summary
        current_row = self.create_statistical_summary_table(ws, current_row, element_data)
        
        # Notes and signature
        current_row = self.create_notes_and_signature_section_optimized(ws, current_row)
        
        return current_row

    def create_measured_values_section(self, ws: Worksheet, start_row: int, element_data: Dict[str, Any]) -> int:
        """Create section with ALL measured values used in charts"""
        
        current_row = start_row
        
        # Get values
        original_values = element_data.get('original_values', [])
        
        if not original_values:
            return current_row
        
        # Section header
        ws.merge_cells(f'A{current_row}:H{current_row}')
        values_header = ws[f'A{current_row}']
        values_header.value = f"MEASURED VALUES (n={len(original_values)})"
        values_header.style = 'section_header'
        ws.row_dimensions[current_row].height = 28
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        
        current_row += 1
        
        # Info text
        ws.merge_cells(f'A{current_row}:H{current_row}')
        info_cell = ws[f'A{current_row}']
        info_cell.value = "All measured values used for chart generation and statistical analysis:"
        info_cell.style = 'param_label'
        ws.row_dimensions[current_row].height = 20
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        
        current_row += 1
        
        # Values in grid format (8 columns)
        values_per_row = 8
        num_rows = (len(original_values) + values_per_row - 1) // values_per_row
        
        for row_idx in range(num_rows):
            ws.row_dimensions[current_row].height = 22
            
            for col_idx in range(values_per_row):
                value_idx = row_idx * values_per_row + col_idx
                
                if value_idx < len(original_values):
                    col_letter = chr(65 + col_idx)  # A, B, C, ...
                    cell = ws[f'{col_letter}{current_row}']
                    cell.value = f"{original_values[value_idx]:.4f}"
                    cell.style = 'data_cell'
                    cell.font = Font(name='Arial', size=8)
                else:
                    # Empty cell
                    col_letter = chr(65 + col_idx)
                    cell = ws[f'{col_letter}{current_row}']
                    cell.value = ""
                    cell.style = 'data_cell'
            
            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
            current_row += 1
        
        # Small spacing
        ws.row_dimensions[current_row].height = 10
        current_row += 1
        
        return current_row

    def add_centered_chart_optimized(self, ws: Worksheet, chart_path: Path, start_row: int) -> int:
        """Add a chart centered in column B"""
        try:
            img = Image(str(chart_path))
            img.width = int(14 * 37.8)
            img.height = int(9 * 37.8)
            ws.add_image(img, f'B{start_row+1}')
            rows_used = (img.height // 18)
            return start_row + rows_used
        except Exception as e:
            self.logger.error(f"Error adding centered chart: {e}")
            return start_row + 14

    def add_side_by_side_chart_optimized(self, ws: Worksheet, chart_path: Path, row: int, position: str) -> int:
        """Add a compact side-by-side chart"""
        try:
            img = Image(str(chart_path))
            img.width = int(10 * 37.8)
            img.height = int(6.5 * 37.8)
            
            if position == 'left':
                left_space = max(0, (4 * 64 - img.width) // 2)
                column_offset = left_space // 64
                anchor_col = chr(65 + min(column_offset, 3))
            else:
                left_space = max(0, (4 * 64 - img.width) // 2)
                column_offset = 4 + (left_space // 64)
                anchor_col = chr(65 + min(column_offset, 7))
            
            ws.add_image(img, f'{anchor_col}{row+1}')
            rows_used = (img.height // 20)
            return row + rows_used
        except Exception as e:
            self.logger.error(f"Error adding side-by-side chart: {e}")
            return row + 12

    def create_notes_and_signature_section_optimized(self, ws: Worksheet, start_row: int) -> int:
        """Create notes and signature section"""
        
        current_row = start_row
        
        # Notes section
        ws.merge_cells(f'A{current_row}:H{current_row}')
        notes_header = ws[f'A{current_row}']
        notes_header.value = "TECHNICAL NOTES & OBSERVATIONS"
        notes_header.style = 'section_header'
        ws.row_dimensions[current_row].height = 26
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        current_row += 1
        
        for i in range(3):
            ws.row_dimensions[current_row + i].height = 22
        
        ws.merge_cells(f'A{current_row}:H{current_row + 2}')
        notes_cell = ws[f'A{current_row}']
        notes_cell.value = ""
        notes_cell.style = 'notes_style'
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row + 2}')
        current_row += 3
        
        ws.row_dimensions[current_row].height = 8
        current_row += 1
        
        # Signature section
        ws.merge_cells(f'A{current_row}:H{current_row}')
        signature_header = ws[f'A{current_row}']
        signature_header.value = "QUALITY APPROVAL"
        signature_header.style = 'section_header'
        ws.row_dimensions[current_row].height = 26
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        current_row += 1
        
        ws.row_dimensions[current_row].height = 22
        ws[f'A{current_row}'].value = "Position:"
        ws[f'A{current_row}'].style = 'param_label'
        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws[f'B{current_row}'].value = "Quality Engineer"
        ws[f'B{current_row}'].style = 'data_cell'
        
        ws[f'E{current_row}'].value = "Date:"
        ws[f'E{current_row}'].style = 'param_label'
        ws.merge_cells(f'F{current_row}:H{current_row}')
        ws[f'F{current_row}'].value = datetime.now().strftime("%d/%m/%Y")
        ws[f'F{current_row}'].style = 'data_cell'
        
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        current_row += 1
        
        ws.row_dimensions[current_row].height = 22
        ws[f'A{current_row}'].value = "Name:"
        ws[f'A{current_row}'].style = 'param_label'
        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws[f'B{current_row}'].style = 'data_cell'
        
        ws[f'E{current_row}'].value = "Signature:"
        ws[f'E{current_row}'].style = 'param_label'
        ws.merge_cells(f'F{current_row}:H{current_row}')
        ws[f'F{current_row}'].style = 'data_cell'
        
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'H{current_row}')
        current_row += 1
        
        return current_row

    def setup_professional_page_formatting(self, ws: Worksheet):
        """Setup professional page formatting"""
        
        column_widths = {
            'A': 18,  'B': 12,  'C': 12,  'D': 12,
            'E': 18,  'F': 12,  'G': 12,  'H': 12
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        ws.page_margins = PageMargins(
            left=0.7, right=0.7, top=0.8, bottom=0.8,
            header=0.3, footer=0.3
        )
        
        ws.page_setup.horizontalCentered = True
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        
        ws.print_options.horizontalCentered = True
        ws.print_options.gridLines = False
        ws.print_options.headings = False
        
        ws.oddFooter.right.text = "&9Page &P of &N"
        ws.evenFooter.right.text = "&9Page &P of &N"

    def create_executive_summary_sheet(self) -> None:
        """Create executive summary cover sheet"""
        
        try:
            summary_ws = self.workbook.create_sheet(title="Executive Summary", index=0)
            
            current_row = 5
            
            # Logo
            logo_path = Path("./assets/images/gui/logo_some.png")
            if logo_path.exists():
                try:
                    img = Image(str(logo_path))
                    img.width = 350
                    img.height = 82
                    summary_ws.add_image(img, 'D1')
                    current_row += 8
                except Exception as e:
                    self.logger.warning(f"Could not add logo: {e}")
                    current_row += 1
            
            # Title
            summary_ws.merge_cells(f'A{current_row}:J{current_row}')
            title_cell = summary_ws[f'A{current_row}']
            title_cell.value = "STATISTICAL PROCESS CONTROL"
            title_cell.font = Font(name='Arial', size=28, bold=True, color=self.COLORS['primary_blue'])
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            summary_ws.row_dimensions[current_row].height = 47
            
            current_row += 1
            
            summary_ws.merge_cells(f'A{current_row}:J{current_row}')
            subtitle_cell = summary_ws[f'A{current_row}']
            subtitle_cell.value = "COMPREHENSIVE ANALYSIS REPORT"
            subtitle_cell.font = Font(name='Arial', size=16, bold=True, color=self.COLORS['secondary_blue'])
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
            summary_ws.row_dimensions[current_row].height = 33
            
            current_row += 1
            
            summary_ws.merge_cells(f'A{current_row}:J{current_row}')
            project_cell = summary_ws[f'A{current_row}']
            project_cell.value = f"{self.client} - {self.ref_project} - {self.batch_number}"
            project_cell.font = Font(name='Arial', size=14, color=self.COLORS['dark_gray'])
            project_cell.alignment = Alignment(horizontal='center', vertical='center')
            summary_ws.row_dimensions[current_row].height = 25
            
            current_row += 4
            
            current_row = self.create_enhanced_capability_visualization(summary_ws, current_row)
            current_row += 3
            
            current_row = self.create_enhanced_project_info(summary_ws, current_row)
            current_row += 3
            
            current_row = self.create_enhanced_elements_summary(summary_ws, current_row)
            
            column_widths = {
                'A': 25, 'B': 12, 'C': 12, 'D': 12, 'E': 12,
                'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 15
            }
            for col, width in column_widths.items():
                summary_ws.column_dimensions[col].width = width
            
            summary_ws.page_setup.horizontalCentered = True
            summary_ws.page_setup.verticalCentered = True
            summary_ws.print_options.horizontalCentered = True
            summary_ws.print_options.verticalCentered = True
            
            self.setup_professional_page_formatting(summary_ws)
            
            self.logger.info("Created executive summary sheet")
            
        except Exception as e:
            self.logger.error(f"Error creating executive summary: {e}")

    def create_enhanced_capability_visualization(self, ws: Worksheet, start_row: int) -> int:
        """Create capability status visualization"""
        current_row = start_row

        ws.merge_cells(f'A{current_row}:J{current_row}')
        vis_header = ws[f'A{current_row}']
        vis_header.value = "PROCESS CAPABILITY OVERVIEW"
        vis_header.style = 'section_header'
        ws.row_dimensions[current_row].height = 35
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
        current_row += 1

        status_counts = {
            "Superior (CPK ≥ 2.0)": 0,
            "Robust (1.67 ≤ CPK < 2.0)": 0,
            "Acceptable (1.33 ≤ CPK < 1.67)": 0,
            "Inadequate (CPK < 1.33)": 0,
            "N/A": 0
        }

        for element_data in self.elements_data.values():
            cpk = element_data.get('cpk', None)
            if isinstance(cpk, (int, float)):
                if cpk >= 2.0:
                    status_counts["Superior (CPK ≥ 2.0)"] += 1
                elif cpk >= 1.67:
                    status_counts["Robust (1.67 ≤ CPK < 2.0)"] += 1
                elif cpk >= 1.33:
                    status_counts["Acceptable (1.33 ≤ CPK < 1.67)"] += 1
                else:
                    status_counts["Inadequate (CPK < 1.33)"] += 1
            else:
                status_counts["N/A"] += 1

        ws.row_dimensions[current_row].height = 30
        headers = ["Capability Status", "Count", "Percentage", "Quality Assessment"]
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'].value = headers[0]
        ws[f'A{current_row}'].style = 'table_header'

        ws.merge_cells(f'C{current_row}:D{current_row}')
        ws[f'C{current_row}'].value = headers[1]
        ws[f'C{current_row}'].style = 'table_header'

        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'].value = headers[2]
        ws[f'E{current_row}'].style = 'table_header'

        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'].value = headers[3]
        ws[f'G{current_row}'].style = 'table_header'

        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
        current_row += 1

        quality_levels = {
            "Superior (CPK ≥ 2.0)": "Superior Process Control - Exceeds Requirements",
            "Robust (1.67 ≤ CPK < 2.0)": "Robust Process - Meets PPAP Standards",
            "Acceptable (1.33 ≤ CPK < 1.67)": "Process Acceptable - Monitor Closely",
            "Inadequate (CPK < 1.33)": "Process Improvement Required - Action Needed",
            "N/A": "Insufficient Data - Additional Analysis Required"
        }

        colors = {
            "Superior (CPK ≥ 2.0)": self.COLORS['excellent_teal'],
            "Robust (1.67 ≤ CPK < 2.0)": self.COLORS['success_green'],
            "Acceptable (1.33 ≤ CPK < 1.67)": self.COLORS['warning_orange'],
            "Inadequate (CPK < 1.33)": self.COLORS['danger_red'],
            "N/A": self.COLORS['medium_gray']
        }

        total_elements = len(self.elements_data)

        for status, count in status_counts.items():
            if count == 0:
                continue

            ws.row_dimensions[current_row].height = 28

            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'].value = status
            ws[f'A{current_row}'].style = 'param_label'

            ws.merge_cells(f'C{current_row}:D{current_row}')
            ws[f'C{current_row}'].value = count
            ws[f'C{current_row}'].style = 'data_cell'
            ws[f'C{current_row}'].font = Font(name='Arial', size=10, bold=True)

            ws.merge_cells(f'E{current_row}:F{current_row}')
            percentage = (count / total_elements) * 100
            ws[f'E{current_row}'].value = f"{percentage:.1f}%"
            ws[f'E{current_row}'].style = 'data_cell'
            ws[f'E{current_row}'].font = Font(name='Arial', size=10, bold=True)

            ws.merge_cells(f'G{current_row}:J{current_row}')
            cell = ws[f'G{current_row}']
            cell.value = quality_levels[status]
            cell.style = 'data_cell'
            cell.fill = PatternFill(start_color=colors[status], end_color=colors[status], fill_type='solid')
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF' if colors[status] == self.COLORS['danger_red'] else '000000')

            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
            current_row += 1

        return current_row

    def create_enhanced_project_info(self, ws: Worksheet, start_row: int) -> int:
        """Create project information section"""
        
        current_row = start_row
        
        ws.merge_cells(f'A{current_row}:J{current_row}')
        info_header = ws[f'A{current_row}']
        info_header.value = "PROJECT INFORMATION"
        info_header.style = 'section_header'
        ws.row_dimensions[current_row].height = 35
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
        
        current_row += 1
        
        date_only_str = datetime.now().strftime("%d/%m/%Y")
        
        first_element_data = next(iter(self.elements_data.values()), {})
        part_description = first_element_data.get('part_description', 'See detailed sheets')
        
        methodology_standards = {
            'cmm': 'ITM-1 CMM Standards',
            'msa': 'AIAG MSA-4 Standards', 
            'manual': 'ITM-2 Manual Standards',
            'optical': 'ITM-3 Optical Standards',
            'laser': 'ITM-4 Laser Standards',
            'gauge': 'ITM-5 Gauge Standards'
        }
        
        methodology = first_element_data.get('methodology', 'cmm')
        standard = methodology_standards.get(str(methodology).lower(), 'PPAP Standards')

        project_info = [
            ["Client:", self.client, "Elements Analyzed:", str(len(self.elements_data))],
            ["Project Reference:", self.ref_project, "Report Generated:", date_only_str],
            ["Batch Number:", self.batch_number, "Part Description:", part_description],
            ["Analysis Date:", date_only_str, "Standard:", standard]
        ]
        
        for info_row in project_info:
            ws.row_dimensions[current_row].height = 28
            
            ws[f'A{current_row}'].value = info_row[0]
            ws[f'A{current_row}'].style = 'param_label'
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'].value = info_row[1]
            ws[f'B{current_row}'].style = 'data_cell'
            ws[f'B{current_row}'].font = Font(name='Arial', size=10)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws.merge_cells(f'E{current_row}:F{current_row}')
            ws[f'E{current_row}'].value = info_row[2]
            ws[f'E{current_row}'].style = 'param_label'
            
            ws.merge_cells(f'G{current_row}:J{current_row}')
            ws[f'G{current_row}'].value = info_row[3]
            ws[f'G{current_row}'].style = 'data_cell'
            ws[f'G{current_row}'].font = Font(name='Arial', size=10)
            ws[f'G{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
            current_row += 1
        
        return current_row

    def create_enhanced_elements_summary(self, ws: Worksheet, start_row: int) -> int:
        """Create elements summary table"""
        
        current_row = start_row
        
        ws.merge_cells(f'A{current_row}:J{current_row}')
        table_header = ws[f'A{current_row}']
        table_header.value = "DETAILED ELEMENTS ANALYSIS SUMMARY"
        table_header.style = 'section_header'
        ws.row_dimensions[current_row].height = 35
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
        
        current_row += 1
        
        headers = [
            "Element", "Cavity", "Mean", "CP", "CPK", 
            "Short σ", "PP", "PPK", "Long σ", "Status"
        ]
        
        ws.row_dimensions[current_row].height = 30
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{current_row}'].value = header
            ws[f'{col}{current_row}'].style = 'table_header'
            ws[f'{col}{current_row}'].font = Font(name='Arial', size=10, bold=True)
        
        self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
        current_row += 1
        
        for element_key, element_data in self.elements_data.items():
            ws.row_dimensions[current_row].height = 25
            
            clean_name = self.extract_element_name(element_key)
            ws[f'A{current_row}'].value = clean_name
            ws[f'A{current_row}'].style = 'data_cell'
            ws[f'A{current_row}'].font = Font(name='Arial', size=9)
            
            ws[f'B{current_row}'].value = element_data.get('cavity', 'N/A')
            ws[f'B{current_row}'].style = 'data_cell'
            
            params = [
                ('mean', 'C', '.4f'),
                ('cp', 'D', '.3f'),
                ('cpk', 'E', '.3f'),
                ('std_short', 'F', '.4f'),
                ('pp', 'G', '.3f'),
                ('ppk', 'H', '.3f'),
                ('std_long', 'I', '.4f')
            ]
            
            for param, col, fmt in params:
                value = element_data.get(param, 'N/A')
                if isinstance(value, (int, float)):
                    formatted_value = f"{value:{fmt}}"
                else:
                    formatted_value = str(value)
                
                ws[f'{col}{current_row}'].value = formatted_value
                ws[f'{col}{current_row}'].style = 'data_cell'
            
            cpk = element_data.get('cpk', 0)
            if isinstance(cpk, (int, float)):
                if cpk >= 2:
                    status = "Excellent"
                    color = self.COLORS['success_green']
                    font_color = self.COLORS['dark_gray']
                elif cpk >= 1.67:
                    status = "Good"
                    color = self.COLORS['excellent_teal']  # FIXED - Turquoise
                    font_color = self.COLORS['dark_gray']  # Dark text on light turquoise
                elif cpk >= 1.33:
                    status = "Acceptable"
                    color = self.COLORS['warning_orange']
                    font_color = self.COLORS['dark_gray']
                else:
                    status = "Inadequate"
                    color = self.COLORS['danger_red']
                    font_color = 'FFFFFF'
            else:
                status = "N/A"
                color = self.COLORS['medium_gray']
                font_color = self.COLORS['dark_gray']
            
            status_cell = ws[f'J{current_row}']
            status_cell.value = status
            status_cell.style = 'data_cell'
            status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            status_cell.font = Font(name='Arial', size=9, bold=True, color=font_color)
            
            self.apply_complete_borders_to_range(ws, f'A{current_row}', f'J{current_row}')
            current_row += 1
        
        return current_row