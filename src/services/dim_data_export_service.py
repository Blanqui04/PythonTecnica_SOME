# src/services/dim_data_export_service.py
import json
import os
import re
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
#from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from src.models.dimensional.dimensional_result import DimensionalResult
from src.database.database_connection import PostgresConn


class DataExportService:
    """Professional dimensional analysis export service for automotive PPAP reports"""

    def __init__(self):
        os.makedirs('logs', exist_ok=True)
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        fh = logging.FileHandler('logs/dimensional.log')    # Create file handler
        fh.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')   # Create formatter
        fh.setFormatter(formatter)
        # Add handler to logger
        if not self.logger.handlers:
            self.logger.addHandler(fh)
        
        self._init_styles()
        
    def _init_styles(self):
        """Initialize Excel styling constants for automotive PPAP standards"""
        # Fonts
        self.HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        self.TITLE_FONT = Font(name='Arial', size=16, bold=True, color='2C3E50')
        self.SUBTITLE_FONT = Font(name='Arial', size=11, bold=True, color='34495E')
        self.DATA_FONT = Font(name='Arial', size=9)
        self.SMALL_FONT = Font(name='Arial', size=8)
        self.NOTES_FONT = Font(name='Arial', size=9)
        self.BASIC_FONT = Font(name='Arial', size=9, color='808080')  # Grey for basic/informative
        self.STATISTICAL_FONT = Font(name='Arial', size=9, bold=True, color='1B4F72')  # Highlight statistical dims
        
        # Fills
        self.HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        self.OK_FILL = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
        self.NOK_FILL = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        self.TED_FILL = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        self.WG_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        self.ALT_ROW_FILL = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        self.STATISTICAL_FILL = PatternFill(start_color='EBF3FD', end_color='EBF3FD', fill_type='solid')  # Light blue for statistical
        
        # Borders - Professional automotive standard
        self.THIN_BORDER = Border(
            left=Side(style='thin', color='D0D3D4'), 
            right=Side(style='thin', color='D0D3D4'),
            top=Side(style='thin', color='D0D3D4'), 
            bottom=Side(style='thin', color='D0D3D4')
        )
        self.MEDIUM_BORDER = Border(
            left=Side(style='medium', color='85929E'), 
            right=Side(style='medium', color='85929E'),
            top=Side(style='medium', color='85929E'), 
            bottom=Side(style='medium', color='85929E')
        )
        self.THICK_BORDER = Border(
            left=Side(style='thick', color='2C3E50'), 
            right=Side(style='thick', color='2C3E50'),
            top=Side(style='thick', color='2C3E50'), 
            bottom=Side(style='thick', color='2C3E50')
        )
        
        # Column separators - Medium borders for logical groupings
        self.LEFT_MEDIUM_BORDER = Border(
            left=Side(style='medium', color='85929E'),
            right=Side(style='thin', color='D0D3D4'),
            top=Side(style='thin', color='D0D3D4'),
            bottom=Side(style='thin', color='D0D3D4')
        )

    def export_dimensional_report(
        self, 
        results: List[DimensionalResult],
        export_dir: str,
        base_filename: str,
        metadata: Dict[str, Any],
        summary_data: Optional[Dict] = None,
        logo_path: Optional[str] = None,
        db_config_path: Optional[str] = None,
        db_key: str = "primary"
    ) -> Dict[str, str]:
        """Main export method for automotive PPAP reports"""
        try:
            os.makedirs(export_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_paths = {}

            # Enhance metadata with database info
            enhanced_metadata = self._metadata(metadata, db_config_path, db_key)
            
            # Sort and group results with proper automotive numbering
            sorted_results = self._sort_results_automotive_standard(results)
            cavity_groups = self._group_results_by_cavity(sorted_results)

            # Create Excel report
            excel_path = os.path.join(export_dir, f"{base_filename}_PPAP_REPORT_{timestamp}.xlsx")
            self._create_professional_excel_report(
                excel_path, 
                cavity_groups,
                enhanced_metadata,
                summary_data,
                logo_path
            )
            export_paths['excel_report'] = excel_path
            # Remove emojis from log messages to avoid UnicodeEncodeError
            self.logger.info(f"Professional Excel PPAP report created: {os.path.basename(excel_path)}")

            # Create comprehensive JSON
            json_path = os.path.join(export_dir, f"{base_filename}_complete_data_{timestamp}.json")
            self._export_comprehensive_json(
                json_path, 
                sorted_results, 
                enhanced_metadata, 
                summary_data,
                cavity_groups
            )
            export_paths['json_data'] = json_path
            self.logger.info(f"Comprehensive JSON exported: {os.path.basename(json_path)}")

            return export_paths

        except Exception as e:
            self.logger.error(f"Export failed: {str(e)}", exc_info=True)
            raise RuntimeError(f"Export failed: {str(e)}") from e

    def _sort_results_automotive_standard(self, results: List[DimensionalResult]) -> List[DimensionalResult]:
        """Sort element_id with automotive industry standard (001, 002, 100, 101, 102, 200, 201.1, 201.2, etc.)"""
        def parse_automotive_id(eid: str) -> Tuple[int, int, float]:
            """Parse automotive element ID for proper sorting"""
            eid = str(eid).strip()
            
            # Remove common prefixes
            cleaned = re.sub(r'^(Nº|No\.?|#)\s*', '', eid, flags=re.IGNORECASE)
            
            try:
                # Handle decimal format (e.g., 201.1, 201.2)
                if '.' in cleaned:
                    parts = cleaned.split('.')
                    main_num = int(parts[0])
                    decimal_part = float('0.' + parts[1]) if len(parts) > 1 else 0
                    return (0, main_num, decimal_part)
                else:
                    # Integer format
                    main_num = int(cleaned)
                    return (0, main_num, 0.0)
                    
            except (ValueError, IndexError):
                # Fallback for non-numeric IDs
                return (1, 999999, 0.0)

        return sorted(results, key=lambda r: parse_automotive_id(r.element_id))

    def _metadata(self, metadata: Dict[str, Any], db_config_path: Optional[str], db_key: str) -> Dict[str, Any]:
        """Enhance metadata with database information and fallbacks"""
        enhanced = {
            'client_name': metadata.get('client_name', ''),
            'project_ref': metadata.get('project_ref', ''),
            'part_number': metadata.get('project_ref', ''),
            'batch_number': metadata.get('batch_number', ''),
            'report_type': metadata.get('report_type', 'PPAP'),
            'drawing_number': '',
            'quotation_number': '',
            'project_leader_name': '',
            'project_leader_title': 'Quality Engineer',
            'quality_facility': '',
            'normative': 'ISO 2768-m',
            'inspector': '',
            'report_date': datetime.now().strftime('%Y-%m-%d'),
            'export_timestamp': datetime.now().isoformat()
        }
        
        # Try to get database metadata
        if db_config_path and os.path.exists(db_config_path):
            try:
                with open(db_config_path, 'r') as f:
                    db_config = json.load(f)
                
                conn_config = db_config.get(db_key, {})
                if conn_config:
                    db_metadata = self._fetch_database_metadata(conn_config, enhanced)
                    enhanced.update(db_metadata)
            except Exception as e:
                self.logger.warning(f"Could not enhance metadata from database: {e}")
        
        return enhanced

    def _fetch_database_metadata(self, conn_config: Dict, base_metadata: Dict) -> Dict[str, Any]:
        """Fetch metadata from database with error handling"""
        try:
            db = PostgresConn(
                host=conn_config["host"],
                database=conn_config["database"],
                user=conn_config["user"],
                password=conn_config["password"],
                port=conn_config.get("port", 5432)
            )

            queries = [
                """SELECT part_number, drawing_number, quotation_number, 
                         project_leader_name, project_leader_title, quality_facility, 
                         normative, inspector
                  FROM project_metadata 
                  WHERE project_ref = %s OR part_number = %s
                  ORDER BY updated_at DESC LIMIT 1""",
                
                """SELECT part_number, drawing_number, quotation_number, 
                         project_leader, '' as title, quality_facility, 
                         normative, inspector
                  FROM projects 
                  WHERE project_ref = %s OR part_number = %s
                  ORDER BY created_at DESC LIMIT 1"""
            ]

            project_ref = base_metadata.get('project_ref', '')
            
            for query in queries:
                try:
                    row = db.fetchone(query, (project_ref, project_ref))
                    if row:
                        return {
                            'part_number': row[0] or base_metadata['part_number'],
                            'drawing_number': row[1] or '',
                            'quotation_number': row[2] or '',
                            'project_leader_name': row[3] or '',
                            'project_leader_title': row[4] or 'Quality Engineer',
                            'quality_facility': row[5] or '',
                            'normative': row[6] or 'ISO 2768-m',
                            'inspector': row[7] or ''
                        }
                except Exception:
                    continue

            db.close()
            
        except Exception as e:
            self.logger.warning(f"Database metadata fetch failed: {e}")
        
        return {}

    def _create_professional_excel_report(
        self,
        filepath: str,
        cavity_groups: Dict[int, List[DimensionalResult]],
        metadata: Dict[str, Any],
        summary_data: Optional[Dict] = None,
        logo_path: Optional[str] = None
    ):
        """Create professional Excel workbook for automotive PPAP"""
        wb = Workbook()
        
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        # Create main report sheet
        main_sheet = wb.create_sheet("PPAP Dimensional Report")
        self._create_main_report_sheet(main_sheet, cavity_groups, metadata, logo_path)
        
        # Create cavity-specific sheets if multiple cavities
        if len(cavity_groups) > 1:
            for cavity_num in sorted(cavity_groups.keys()):
                cavity_sheet = wb.create_sheet(f"Cavity {cavity_num}")
                self._create_cavity_specific_sheet(
                    cavity_sheet, 
                    cavity_groups[cavity_num], 
                    cavity_num, 
                    metadata, 
                    logo_path
                )
        
        # Create summary sheet
        if summary_data:
            summary_sheet = wb.create_sheet("Analysis Summary")
            self._create_summary_sheet(summary_sheet, cavity_groups, metadata, summary_data)
        
        # Create metadata sheet
        metadata_sheet = wb.create_sheet("Report Metadata")
        self._create_metadata_sheet(metadata_sheet, metadata, summary_data)
        
        # Set print settings for all sheets
        for sheet in wb.worksheets:
            self._set_print_settings(sheet)
        
        wb.save(filepath)

    def _create_main_report_sheet(
        self,
        ws: Worksheet,
        cavity_groups: Dict[int, List[DimensionalResult]],
        metadata: Dict[str, Any],
        logo_path: Optional[str] = None
    ):
        """Create main PPAP report sheet with professional automotive header"""
        current_row = 1
        
        # Add professional header
        current_row = self._add_ppap_header(ws, metadata, current_row, logo_path)
        current_row += 2
        
        # Add dimensional data table with automotive specifications
        all_results = []
        for cavity_results in cavity_groups.values():
            all_results.extend(cavity_results)
        
        current_row = self._add_automotive_dimensional_table(ws, all_results, current_row)
        current_row += 3
        
        # Add professional footer
        self._add_ppap_footer(ws, metadata, current_row)
        
        # Apply professional formatting
        self._apply_automotive_formatting(ws)

    def _add_automotive_dimensional_table(self, ws: Worksheet, results: List[DimensionalResult], start_row: int) -> int:
        """Add automotive industry standard dimensional data table with improved header aesthetics"""
        current_row = start_row
        
        # Enhanced headers with better formatting and spacing
        headers = [
            'Element\nID', 'Description', 'Measuring\nInstrument', 'M1', 'M2', 'M3', 'M4', 'M5',
            'Min\nValue', 'Max\nValue', 'Mean\nValue', 'Std\nDev', 'Pp', 'Ppk', 'Status'
        ]
        
        # Write headers with enhanced automotive styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Apply enhanced borders with logical groupings
            if col == 3:  # After Measuring Instrument
                cell.border = Border(
                    left=Side(style='thin', color='D0D3D4'),
                    right=Side(style='medium', color='85929E'),
                    top=Side(style='thick', color='2C3E50'),
                    bottom=Side(style='thick', color='2C3E50')
                )
            elif col == 8:  # After M5
                cell.border = Border(
                    left=Side(style='thin', color='D0D3D4'),
                    right=Side(style='medium', color='85929E'),
                    top=Side(style='thick', color='2C3E50'),
                    bottom=Side(style='thick', color='2C3E50')
                )
            elif col == 11:  # After Mean Value
                cell.border = Border(
                    left=Side(style='thin', color='D0D3D4'),
                    right=Side(style='medium', color='85929E'),
                    top=Side(style='thick', color='2C3E50'),
                    bottom=Side(style='thick', color='2C3E50')
                )
            elif col == 14:  # After Ppk
                cell.border = Border(
                    left=Side(style='thin', color='D0D3D4'),
                    right=Side(style='medium', color='85929E'),
                    top=Side(style='thick', color='2C3E50'),
                    bottom=Side(style='thick', color='2C3E50')
                )
            else:
                cell.border = self.THICK_BORDER
        
        # Increased header row height for better readability
        ws.row_dimensions[current_row].height = 40
        current_row += 1
        
        # Write data rows with automotive formatting
        for row_idx, result in enumerate(results):
            self._add_automotive_data_row(ws, current_row, result, len(headers), row_idx)
            current_row += 1
        
        # Set automotive column widths
        self._set_automotive_column_widths(ws)
        
        return current_row

    def _add_automotive_data_row(self, ws: Worksheet, row: int, result: DimensionalResult, num_columns: int, row_index: int):
        """Add data row with automotive industry formatting standards - using pre-calculated values from DimensionalResult"""

        # --- FIX: Robust evaluation_type detection for notes ---
        # Try to get evaluation_type from result, fallback to checking description or class for 'note'
        evaluation_type = (
            getattr(result, 'evaluation_type', None)
            or getattr(result, 'feature_type', None)
            or ''
        )
        evaluation_type = str(evaluation_type).strip().lower()

        # Fallback: if evaluation_type is empty, check if description or class hints at note
        is_notes = False
        if evaluation_type in ['note', 'notes']:
            is_notes = True
        elif not evaluation_type:
            # Try to infer from description or class
            desc = (getattr(result, 'description', '') or '').lower()
            classe = (getattr(result, 'classe', '') or '').lower()
            if 'note' in desc or 'nota' in desc or classe == 'note':
                is_notes = True

        dimension_class = getattr(result, 'classe', '') or ''
        is_basic_info = evaluation_type in ['basic', 'informative']
        is_statistical = dimension_class.upper() in ['CC', 'SC', 'IC'] and not is_notes

        measurements = result.measurements if result.measurements else []
        has_measurements = bool(measurements and any(m is not None for m in measurements))

        print(f"[TABLE] Adding row {row} for element_id={result.element_id} | eval_type={evaluation_type} | class={dimension_class} | is_note={is_notes}")

        # If there are no measurements, force all stats/measurements/status to empty/TO CHECK
        if not has_measurements or is_notes:
            mean_val = None
            std_val = None
            min_val = None
            max_val = None
            pp_val = None
            ppk_val = None
            measurements_out = [""] * 5
            status = "TO CHECK"
        else:
            mean_val = result.mean if hasattr(result, 'mean') and result.mean is not None else None
            min_val = min(measurements) if measurements else None
            max_val = max(measurements) if measurements else None
            measurements_out = [self._format_measurement(m) for m in (measurements + [None] * 5)[:5]]
            status = result.status.value if hasattr(result.status, "value") else str(result.status)
            # Only show std, pp, ppk if statistical
            if is_statistical:
                std_val = result.std_dev if hasattr(result, 'std_dev') and result.std_dev is not None else None
                pp_val = result.pp if hasattr(result, 'pp') and result.pp is not None else None
                ppk_val = result.ppk if hasattr(result, 'ppk') and result.ppk is not None else None
            else:
                std_val = None
                pp_val = None
                ppk_val = None

        # Prepare data using pre-calculated values
        data = [
            result.element_id,
            result.description or '',
            self._format_instrument(result.measuring_instrument),
            *measurements_out,
            "" if not has_measurements else self._format_number(min_val, 2),
            "" if not has_measurements else self._format_number(max_val, 2),
            "" if not has_measurements else self._format_number(mean_val, 2),
            "" if not has_measurements or not is_statistical else self._format_number(std_val, 2),
            "" if not has_measurements or not is_statistical else self._format_number(pp_val, 2),
            "" if not has_measurements or not is_statistical else self._format_number(ppk_val, 2),
            status
        ]

        is_alternate = row_index % 2 == 1

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)

            # Apply font based on dimension type
            if not has_measurements or is_notes:
                cell.font = self.NOTES_FONT
            elif is_basic_info:
                cell.font = self.BASIC_FONT
            elif is_statistical:
                cell.font = self.STATISTICAL_FONT
            else:
                cell.font = self.DATA_FONT

            # Alignment
            if col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 2:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif col == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif 4 <= col <= 8:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif 9 <= col <= 14:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 15:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Background fills
            if col == 15:
                self._apply_status_formatting(cell, self._normalize_status(status))
            elif is_statistical:
                cell.fill = self.STATISTICAL_FILL
            elif is_alternate:
                cell.fill = self.ALT_ROW_FILL

            # Borders with grouping separators
            if col == 4:
                cell.border = self.LEFT_MEDIUM_BORDER
            elif col == 9:
                cell.border = self.LEFT_MEDIUM_BORDER
            elif col == 12:
                cell.border = self.LEFT_MEDIUM_BORDER
            elif col == 15:
                cell.border = self.LEFT_MEDIUM_BORDER
            else:
                cell.border = self.THIN_BORDER

        if not has_measurements or is_notes:
            ws.row_dimensions[row].height = 40

    def _normalize_status(self, status) -> str:
        """Normalize status values for consistency - handles both enum and string values"""
        # Handle enum values
        if hasattr(status, 'value'):
            status_str = status.value
        else:
            status_str = str(status)
        
        status_str = status_str.strip().upper()
        
        status_map = {
            'GOOD': 'OK', 'PASS': 'OK', 'OK': 'OK', 'PASSED': 'OK',
            'BAD': 'NOK', 'FAIL': 'NOK', 'NOK': 'NOK', 'NG': 'NOK', 'FAILED': 'NOK',
            'WARNING': 'WARNING', 'WARN': 'WARNING',
            'T.E.D.': 'T.E.D', 'TED': 'T.E.D', 'T.E.D': 'T.E.D', 'BASIC': 'T.E.D', 'INFORMATIVE': 'T.E.D',
            'TO CHECK': 'TO CHECK', 'TO_CHECK': 'TO CHECK', 'CHECK': 'TO CHECK'
        }
        
        return status_map.get(status_str, 'WARNING')
    
    def _format_measurement(self, value) -> str:
        """Format measurement values to 2 decimals, left aligned"""
        try:
            if value is None:
                return ''
            num = float(value)
            return f"{num:.2f}"
        except (ValueError, TypeError):
            return str(value) if value is not None else ''

    def _format_number(self, value, decimals: int = 2) -> str:
        """Format numbers consistently to specified decimals"""
        try:
            if value is None:
                return ''
            num = float(value)
            return f"{num:.{decimals}f}"
        except (ValueError, TypeError):
            return str(value) if value is not None else ''

    def _format_instrument(self, instrument: Optional[str]) -> str:
        """Format instrument name consistently"""
        if not instrument:
            return ''
        i = str(instrument).strip()
        return i.capitalize() if i.lower() == 'visual' else i

    def _apply_status_formatting(self, cell, status: str):
        """Apply status-based formatting for automotive standards"""
        status_formats = {
            'OK': (self.OK_FILL, Font(name='Arial', size=9, bold=True, color='27AE60')),
            'NOK': (self.NOK_FILL, Font(name='Arial', size=9, bold=True, color='E74C3C')),
            'T.E.D': (self.TED_FILL, Font(name='Arial', size=9, bold=True, color='3498DB')),
            'WARNING': (self.WG_FILL, Font(name='Arial', size=9, bold=True, color='856404')),
            'TO CHECK': (PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'), 
                        Font(name='Arial', size=9, bold=True, color='D6B656'))
        }
        
        if status in status_formats:
            fill, font = status_formats[status]
            cell.fill = fill
            cell.font = font

    def _add_ppap_header(
        self, 
        ws: Worksheet, 
        metadata: Dict[str, Any], 
        start_row: int, 
        logo_path: Optional[str] = None
    ) -> int:
        """Add professional PPAP header for automotive industry"""
        current_row = start_row
        
        # Logo and title row
        if logo_path and os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                img.height = 50
                img.width = 100
                ws.add_image(img, f'A{current_row}')
            except Exception as e:
                self.logger.warning(f"Could not add logo: {e}")
        
        # Main title
        ws.merge_cells(f'E{current_row}:L{current_row}')
        title_cell = ws.cell(row=current_row, column=5, value="DIMENSIONAL ANALYSIS REPORT")
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Report type
        ws.merge_cells(f'M{current_row}:O{current_row}')
        type_cell = ws.cell(row=current_row, column=13, value=f"{metadata.get('report_type', 'PPAP')} REPORT")
        type_cell.font = Font(name='Arial', size=12, bold=True, color='E74C3C')
        type_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 3
        
        # Information table - automotive specific
        info_data = [
            ('Client / Supplier:', metadata.get('client_name', ''), 'Part Number:', metadata.get('part_number', '')),
            ('Drawing Number:', metadata.get('drawing_number', ''), 'Batch Number:', metadata.get('batch_number', '')),
            ('Project Reference:', metadata.get('project_ref', ''), 'Report Date:', metadata.get('report_date', '')),
            ('Quality Facility:', metadata.get('quality_facility', ''), 'Normative:', metadata.get('normative', 'ISO 2768-m')),
            ('Project Leader:', metadata.get('project_leader_name', ''), 'Inspector:', metadata.get('inspector', ''))
        ]
        
        for left_label, left_value, right_label, right_value in info_data:
            # Left side
            ws.cell(row=current_row, column=1, value=left_label).font = self.SUBTITLE_FONT
            ws.merge_cells(f'B{current_row}:G{current_row}')
            ws.cell(row=current_row, column=2, value=left_value).font = self.DATA_FONT
            
            # Right side  
            ws.cell(row=current_row, column=9, value=right_label).font = self.SUBTITLE_FONT
            ws.merge_cells(f'J{current_row}:O{current_row}')
            ws.cell(row=current_row, column=10, value=right_value).font = self.DATA_FONT
            
            # Apply borders
            for col in range(1, 16):
                ws.cell(row=current_row, column=col).border = self.THIN_BORDER
            
            current_row += 1
        
        return current_row

    def _add_ppap_footer(self, ws: Worksheet, metadata: Dict[str, Any], start_row: int):
        """Add professional PPAP footer with signatures for automotive compliance"""
        current_row = start_row
        
        # Notes section
        ws.merge_cells(f'A{current_row}:O{current_row}')
        notes_cell = ws.cell(row=current_row, column=1, value="NOTES / OBSERVATIONS:")
        notes_cell.font = self.SUBTITLE_FONT
        notes_cell.border = self.THIN_BORDER
        current_row += 1
        
        # Notes box (4 rows)
        for i in range(4):
            for col in range(1, 16):
                cell = ws.cell(row=current_row, column=col, value='')
                cell.border = self.THIN_BORDER
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        current_row += 1
        
        # Signature section
        # Project Leader
        ws.cell(row=current_row, column=2, value="PROJECT LEADER:").font = self.SUBTITLE_FONT
        ws.cell(row=current_row + 1, column=2, value=metadata.get('project_leader_name', '')).font = self.DATA_FONT
        ws.cell(row=current_row + 2, column=2, value=metadata.get('project_leader_title', 'Quality Engineer')).font = self.SMALL_FONT
        ws.cell(row=current_row + 3, column=2, value="Signature: _________________").font = self.DATA_FONT
        
        # Date
        ws.cell(row=current_row, column=10, value="DATE:").font = self.SUBTITLE_FONT
        ws.cell(row=current_row + 1, column=10, value=metadata.get('report_date', '')).font = self.DATA_FONT

    def _set_automotive_column_widths(self, ws: Worksheet):
        """Set column widths optimized for automotive PPAP reports with enhanced headers"""
        column_widths = {
            'A': 12,   # Element ID
            'B': 35,   # Description (wider for better readability)
            'C': 16,   # Measuring Instrument (slightly wider)
            'D': 8,    # M1
            'E': 8,    # M2
            'F': 8,    # M3
            'G': 8,    # M4
            'H': 8,    # M5
            'I': 9,    # Min Value (slightly wider)
            'J': 9,    # Max Value (slightly wider)
            'K': 9,    # Mean Value (slightly wider)
            'L': 8,    # Std Dev
            'M': 8,    # Pp
            'N': 8,    # Ppk
            'O': 12    # Status (wider for "TO CHECK")
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def _apply_automotive_formatting(self, ws: Worksheet):
        """Apply professional automotive industry formatting"""
        # Set row heights for readability
        for row in ws.iter_rows():
            if row[0].row > 1:  # Skip header row
                ws.row_dimensions[row[0].row].height = 25

    def _create_cavity_specific_sheet(
        self,
        ws: Worksheet,
        cavity_results: List[DimensionalResult],
        cavity_num: int,
        metadata: Dict[str, Any],
        logo_path: Optional[str] = None
    ):
        """Create cavity-specific sheet with automotive formatting"""
        cavity_metadata = {**metadata, 'title_suffix': f" - Cavity {cavity_num}"}
        
        current_row = 1
        current_row = self._add_ppap_header(ws, cavity_metadata, current_row, logo_path)
        current_row += 2
        current_row = self._add_automotive_dimensional_table(ws, cavity_results, current_row)
        current_row += 3
        self._add_ppap_footer(ws, metadata, current_row)
        self._apply_automotive_formatting(ws)

    def _create_summary_sheet(
        self,
        ws: Worksheet,
        cavity_groups: Dict[int, List[DimensionalResult]],
        metadata: Dict[str, Any],
        summary_data: Dict[str, Any]
    ):
        """Create professional summary sheet with automotive metrics - using pre-calculated values"""
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title = ws.cell(row=current_row, column=1, value="DIMENSIONAL ANALYSIS SUMMARY")
        title.font = self.TITLE_FONT
        title.alignment = Alignment(horizontal='center')
        current_row += 3
        
        # Calculate automotive-specific statistics using pre-calculated statuses
        all_results = []
        for cavity_results in cavity_groups.values():
            all_results.extend(cavity_results)
        
        total_dims = len(all_results)
        
        # Use pre-calculated statuses from DimensionalResult objects
        statuses = [self._normalize_status(r.status) for r in all_results]
        ok_count = statuses.count('OK')
        nok_count = statuses.count('NOK')
        ted_count = statuses.count('T.E.D')
        warning_count = statuses.count('WARNING')
        
        # Count statistical dimensions using has_classification method or classe attribute
        statistical_count = sum(1 for r in all_results 
                              if hasattr(r, 'has_classification') and r.has_classification() 
                              or getattr(r, 'classe', '').upper() in ['CC', 'SC', 'IC'])
        
        success_rate = (ok_count / (ok_count + nok_count)) * 100 if (ok_count + nok_count) > 0 else 0
        
        # Overall statistics
        stats = [
            ('Total Dimensions:', total_dims),
            ('Passed (OK):', ok_count),
            ('Failed (NOK):', nok_count),
            ('Basic/Info (T.E.D):', ted_count),
            ('Warnings:', warning_count),
            ('Statistical Dimensions:', statistical_count),
            ('Success Rate:', f"{success_rate:.1f}%"),
            ('Analysis Date:', metadata.get('report_date', ''))
        ]
        
        for label, value in stats:
            ws.cell(row=current_row, column=1, value=label).font = self.SUBTITLE_FONT
            ws.cell(row=current_row, column=3, value=value).font = self.DATA_FONT
            current_row += 1
        
        # Cavity breakdown if multiple cavities
        if len(cavity_groups) > 1:
            current_row += 2
            ws.cell(row=current_row, column=1, value="CAVITY BREAKDOWN:").font = self.SUBTITLE_FONT
            current_row += 1
            
            for cavity_num in sorted(cavity_groups.keys()):
                cavity_results = cavity_groups[cavity_num]
                cavity_statuses = [self._normalize_status(r.status) for r in cavity_results]
                cavity_ok = cavity_statuses.count('OK')
                cavity_nok = cavity_statuses.count('NOK')
                cavity_rate = (cavity_ok / (cavity_ok + cavity_nok)) * 100 if (cavity_ok + cavity_nok) > 0 else 0
                
                ws.cell(row=current_row, column=1, value=f"Cavity {cavity_num}:").font = self.DATA_FONT
                ws.cell(row=current_row, column=3, value=f"{cavity_ok}/{cavity_ok + cavity_nok} passed ({cavity_rate:.1f}%)").font = self.DATA_FONT
                current_row += 1

    def _create_metadata_sheet(self, ws: Worksheet, metadata: Dict[str, Any], summary_data: Optional[Dict]):
        """Create metadata sheet with automotive compliance information"""
        current_row = 1
        
        ws.merge_cells(f'A{current_row}:B{current_row}')
        title = ws.cell(row=current_row, column=1, value="REPORT METADATA")
        title.font = self.TITLE_FONT
        current_row += 3
        
        # Metadata sections for automotive compliance
        sections = [
            ("Project Information", [
                ("Client Name", metadata.get('client_name', '')),
                ("Project Reference", metadata.get('project_ref', '')),
                ("Part Number", metadata.get('part_number', '')),
                ("Drawing Number", metadata.get('drawing_number', '')),
                ("Batch Number", metadata.get('batch_number', '')),
                ("Report Type", metadata.get('report_type', 'PPAP'))
            ]),
            ("Quality Information", [
                ("Project Leader", metadata.get('project_leader_name', '')),
                ("Project Leader Title", metadata.get('project_leader_title', '')),
                ("Quality Facility", metadata.get('quality_facility', '')),
                ("Inspector", metadata.get('inspector', '')),
                ("Normative Standard", metadata.get('normative', 'ISO 2768-m'))
            ]),
            ("Export Information", [
                ("Export Date", metadata.get('report_date', '')),
                ("Export Timestamp", metadata.get('export_timestamp', '')),
                ("Software Version", "2.0 Automotive"),
                ("Compliance", "Automotive PPAP Standards")
            ])
        ]
        
        for section_title, section_data in sections:
            ws.cell(row=current_row, column=1, value=section_title).font = self.SUBTITLE_FONT
            current_row += 1
            
            for label, value in section_data:
                ws.cell(row=current_row, column=1, value=label).font = self.DATA_FONT
                ws.cell(row=current_row, column=2, value=value).font = self.DATA_FONT
                current_row += 1
            
            current_row += 1

    def _export_comprehensive_json(
        self,
        filepath: str,
        results: List[DimensionalResult],
        metadata: Dict[str, Any],
        summary_data: Optional[Dict] = None,
        cavity_groups: Optional[Dict] = None
    ):
        """Export comprehensive JSON data with automotive specifications - using pre-calculated values"""
        export_data = {
            "metadata": metadata,
            "results": [self._result_to_dict(r) for r in results],
            "cavity_groups": {str(k): [self._result_to_dict(r) for r in v] for k, v in (cavity_groups or {}).items()},
            "summary": summary_data or {},
            "statistics": self._calculate_overall_statistics(results),
            "automotive_compliance": {
                "ppap_standard": True,
                "statistical_dimensions": len([r for r in results 
                                             if hasattr(r, 'has_classification') and r.has_classification() 
                                             or getattr(r, 'classe', '').upper() in ['CC', 'SC', 'IC']]),
                "export_format_version": "2.0_automotive"
            },
            "export_timestamp": datetime.now().isoformat(),
            "version": "2.0"
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)

    def _calculate_overall_statistics(self, results: List[DimensionalResult]) -> Dict[str, Any]:
        """Calculate overall statistics for automotive compliance - using pre-calculated values"""
        if not results:
            return {}
        
        # Use pre-calculated statuses from DimensionalResult objects
        statuses = [self._normalize_status(r.status) for r in results]
        
        # Count statistical dimensions using pre-calculated classification
        statistical_dims = [r for r in results 
                          if hasattr(r, 'has_classification') and r.has_classification() 
                          or getattr(r, 'classe', '').upper() in ['CC', 'SC', 'IC']]
        
        ok_count = statuses.count('OK')
        nok_count = statuses.count('NOK')
        
        return {
            "total_dimensions": len(results),
            "passed": ok_count,
            "failed": nok_count,
            "basic_info": statuses.count('T.E.D'),
            "warnings": statuses.count('WARNING'),
            "statistical_dimensions": len(statistical_dims),
            "success_rate": (ok_count / (ok_count + nok_count)) * 100 if (ok_count + nok_count) > 0 else 0,
            "automotive_compliance": True
        }

    def _group_results_by_cavity(self, results: List[DimensionalResult]) -> Dict[int, List[DimensionalResult]]:
        """Group results by cavity number with proper automotive handling"""
        cavity_groups = {}
        
        for result in results:
            # Use the cavity attribute from DimensionalResult
            cavity = getattr(result, 'cavity', '1')
            
            # Handle string cavity values and convert to int
            try:
                cavity_num = int(str(cavity).strip()) if cavity else 1
            except (ValueError, TypeError):
                cavity_num = 1  # Default to cavity 1
            
            if cavity_num not in cavity_groups:
                cavity_groups[cavity_num] = []
            cavity_groups[cavity_num].append(result)
        
        return cavity_groups

    def _result_to_dict(self, result: DimensionalResult) -> Dict[str, Any]:
        """Convert result to dictionary with automotive-specific data - using pre-calculated values"""

        # Use pre-calculated values from DimensionalResult
        measurements = list(result.measurements) if result.measurements else []

        stats = {
            'min': min(measurements) if measurements else None,
            'max': max(measurements) if measurements else None,
            'mean': getattr(result, 'mean', None),
            'std': getattr(result, 'std_dev', None)
        }

        is_statistical = (hasattr(result, 'has_classification') and result.has_classification()) or \
                        getattr(result, 'classe', '').upper() in ['CC', 'SC', 'IC']

        pp_ppk = {
            'pp': getattr(result, 'pp', None) if is_statistical else None,
            'ppk': getattr(result, 'ppk', None) if is_statistical else None
        }

        # Use status as is (do not normalize for JSON export)
        status = result.status.value if hasattr(result.status, "value") else str(result.status)

        # For notes, ensure measurements/stats are empty and status is TO CHECK
        evaluation_type = (getattr(result, 'evaluation_type', '') or '').strip().lower()
        if evaluation_type in ['note', 'notes']:
            measurements = []
            stats = {'min': None, 'max': None, 'mean': None, 'std': None}
            pp_ppk = {'pp': None, 'ppk': None}
            status = "TO CHECK"

        return {
            "element_id": result.element_id,
            "description": getattr(result, 'description', ''),
            "batch": getattr(result, 'batch', ''),
            "cavity": getattr(result, 'cavity', '1'),
            "classe": getattr(result, 'classe', ''),
            "nominal": getattr(result, 'nominal', None),
            "lower_tolerance": getattr(result, 'lower_tolerance', None),
            "upper_tolerance": getattr(result, 'upper_tolerance', None),
            "evaluation_type": getattr(result, 'evaluation_type', ''),
            "measuring_instrument": getattr(result, 'measuring_instrument', 'ScanBox'),
            "measurements": measurements,
            "deviation": list(getattr(result, 'deviation', [])) if hasattr(result, 'deviation') else [],
            "statistics": stats,
            "pp_ppk": pp_ppk,
            "status": status,
            "out_of_spec_count": getattr(result, 'out_of_spec_count', 0),
            "gdt_flags": getattr(result, 'gdt_flags', {}),
            "warnings": list(getattr(result, 'warnings', [])) if hasattr(result, 'warnings') else [],
            "is_statistical": is_statistical,
            "automotive_compliant": True
        }

    def _set_print_settings(self, ws: Worksheet):
        """Set professional print settings for automotive reports"""
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = True
        ws.page_setup.fitToWidth = 1
        
        # Set margins for automotive standard
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.75, bottom=0.75, 
            header=0.3, footer=0.3
        )
        
        # Print titles (repeat first row on each page)
        ws.print_title_rows = '1:1'

    def generate_export_summary(self, export_paths: Dict[str, str], metadata: Dict[str, Any]) -> str:
        """Generate professional export summary for automotive clients"""
        try:
            summary_lines = [
                "AUTOMOTIVE DIMENSIONAL ANALYSIS EXPORT SUMMARY",
                "=" * 60,
                "",
                "PROJECT INFORMATION:",
                f"  Client: {metadata.get('client_name', 'N/A')}",
                f"  Project Reference: {metadata.get('project_ref', 'N/A')}",
                f"  Part Number: {metadata.get('part_number', 'N/A')}",
                f"  Batch Number: {metadata.get('batch_number', 'N/A')}",
                f"  Report Type: {metadata.get('report_type', 'PPAP')}",
                "",
                "QUALITY INFORMATION:",
                f"  Project Leader: {metadata.get('project_leader_name', 'N/A')}",
                f"  Quality Facility: {metadata.get('quality_facility', 'N/A')}",
                f"  Normative Standard: {metadata.get('normative', 'ISO 2768-m')}",
                "",
                "EXPORTED FILES:",
                f"  • Professional Excel Report: {os.path.basename(export_paths.get('excel_report', 'Not generated'))}",
                f"  • Comprehensive JSON Data: {os.path.basename(export_paths.get('json_data', 'Not generated'))}",
                "",
                f"Export completed successfully at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                "",
                "AUTOMOTIVE PPAP COMPLIANCE FEATURES:",
                "  ✓ Professional automotive-standard formatting",
                "  ✓ Proper dimensional sorting (001, 002, 100, 201.1, 201.2, etc.)",
                "  ✓ Statistical analysis (Pp/Ppk) using pre-calculated values",
                "  ✓ Status validation using pre-calculated DimensionalResult status",
                "  ✓ Notes dimensions with proper text alignment",
                "  ✓ Basic/Informative dimensions marked as T.E.D",
                "  ✓ Logical column groupings with medium borders",
                "  ✓ Alternating row colors for readability",
                "  ✓ Cavity-specific analysis and sorting",
                "  ✓ Professional PPAP header and signature areas",
                "  ✓ Uses pre-calculated values from DimensionalResult objects",
                "",
                "Ready for submission to automotive clients:",
                "Autoliv • ZF • Brose • Sofegi • BMW • Mercedes • Audi • VW",
                "=" * 60
            ]
            
            return "\n".join(summary_lines)
            
        except Exception as e:
            self.logger.error(f"Could not generate export summary: {str(e)}")
            return f"Export completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\nFiles: {len(export_paths)} generated"