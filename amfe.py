import os
import zipfile
import openpyxl
import warnings
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
import openpyxl.styles
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------- Variables de preàmbul ---------- #
NOM_CARPETA_AMFES = r'\\Server\some\1002-SPS\3.5 PROCESS ASSURANCE_P-FMEA\4 Draft\00_AMFES en curs'
CARPETA_AMFE = r'\03_AMFE'

class AmfeManager:
    def __init__(self, base_folder=NOM_CARPETA_AMFES):
        self.base_folder = base_folder
        self.available_processes = self.scan_available_processes()

    def scan_available_processes(self):
        try:
            dirs = [
                d for d in os.listdir(self.base_folder)
                if os.path.isdir(os.path.join(self.base_folder, d))
            ]
            return dirs
        except Exception as e:
            print(f"Error escanejant processos: {e}")
            return []

    def get_amfe_paths(self, selected_processes):
        amfe_paths = []
        for proc in selected_processes:
            proc_folder = os.path.join(self.base_folder, proc)
            amfe_folder = os.path.join(proc_folder, CARPETA_AMFE.strip("\\/"))
            if os.path.exists(amfe_folder):
                amfe_paths.append(amfe_folder)
        return amfe_paths
    
    def get_amfe_excels(self, paths):
        """
        Retorna una llista de fitxers Excel vàlids, ignorant fitxers temporals
        """
        amfe_excels = []
        for path in paths:
            try:
                # Ignore temporary Excel files (starting with '~$')
                files = [f for f in os.listdir(path) 
                         if f.endswith('.xlsx') and not f.startswith('~$')]
                for file in files:
                    full_path = os.path.join(path, file)
                    amfe_excels.append(full_path)
            except Exception as e:
                print(f"Error llegint fitxers a {path}: {e}")
        return amfe_excels

    def is_row_empty(self, row):
        """Check if all cells in a row are empty."""
        return all(cell is None or cell == '' for cell in row)

    def is_valid_excel(self, file_path):
        """Check if file is a valid Excel file"""
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return False
            
        if not zipfile.is_zipfile(file_path):
            print(f"Not a valid Excel file (not a ZIP archive): {file_path}")
            return False
            
        try:
            with zipfile.ZipFile(file_path) as zf:
                return 'xl/workbook.xml' in zf.namelist()
        except zipfile.BadZipFile:
            print(f"Corrupted Excel file (BadZipFile): {file_path}")
        except Exception as e:
            print(f"Error checking Excel file: {file_path} - {str(e)}")
        return False

    def read_amfe_data(self, file_path, include_header):
        """
        Llegeix dades d'un fitxer AMFE amb gestió d'errors.
        Retorna: llista de files (cada fila és una llista de valors)
        """
        # Validate file before processing
        if not self.is_valid_excel(file_path):
            print(f"Skipping invalid file: {file_path}")
            return []
            
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            print(f"Error loading workbook {file_path}: {str(e)}")
            return []

        try:
            sheet_name = 'P-FMEA_AIAG-VDA'
            if sheet_name not in wb.sheetnames:
                print(f"Advertència: Full '{sheet_name}' no trobat a {file_path}")
                return []
            
            sheet = wb[sheet_name]
            data_rows = []
            max_row = sheet.max_row
            columns_to_keep = list(range(1, 19)) + list(range(22, 35))  # A-R i V-AH

            # Processar capçalera si cal
            if include_header:
                # Fila 1
                row1 = [sheet.cell(row=1, column=c).value for c in range(1, 35)]
                row1_trimmed = [row1[i-1] for i in columns_to_keep]
                data_rows.append(row1_trimmed)
                
                # Fila 7
                row7 = [sheet.cell(row=7, column=c).value for c in range(1, 35)]
                row7_trimmed = [row7[i-1] for i in columns_to_keep]
                data_rows.append(row7_trimmed)
                
                # Fila 8
                row8 = [sheet.cell(row=8, column=c).value for c in range(1, 35)]
                row8_trimmed = [row8[i-1] for i in columns_to_keep]
                data_rows.append(row8_trimmed)

            # Processar dades (a partir de fila 10)
            current_row = 10
            while current_row <= max_row:
                row_data = [sheet.cell(row=current_row, column=c).value for c in range(1, 35)]
                
                # Retalla les columnes no desitjades (S, T, U)
                trimmed_row = [row_data[i-1] for i in columns_to_keep]
                
                # Aturem al trobar la primera fila buida
                if self.is_row_empty(trimmed_row):
                    break
                
                data_rows.append(trimmed_row)
                current_row += 1
            
            return data_rows
            
        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
            return []
        finally:
            wb.close()

    def count_risk_levels(self, all_rows):
        """
        Compta els nivells de risc (H, M, L) per a les columnes P (abans) i AG (després)
        """
        before_counts = {'H': 0, 'M': 0, 'L': 0}
        after_counts = {'H': 0, 'M': 0, 'L': 0}
        
        # Índexs de les columnes (0-based)
        col_before = 16  # Columna Q (abans)
        col_after = 29   # Columna AG (després)
        
        # Comencem des de la fila 4 (ignorant les tres primeres files de capçalera)
        for row in all_rows[3:]:
            # Saltar files buides o separadores
            if row is None or all(cell is None for cell in row):
                continue
                
            # Verificar si la fila té suficients columnes
            if len(row) > col_before:
                # Handle NaN values
                value = row[col_before]
                if isinstance(value, float) and value != value:  # Check for NaN
                    risk_before = ''
                else:
                    risk_before = str(value).strip().upper() if value else ''
                    
                if risk_before in ['H', 'M', 'L']:
                    before_counts[risk_before] += 1
                    
            if len(row) > col_after:
                # Handle NaN values
                value = row[col_after]
                if isinstance(value, float) and value != value:  # Check for NaN
                    risk_after = ''
                else:
                    risk_after = str(value).strip().upper() if value else ''
                    
                if risk_after in ['H', 'M', 'L']:
                    after_counts[risk_after] += 1
                    
        return before_counts, after_counts
    
    
    def get_risk_summary_data(self, filepath, show='before'):
        """
        Correctly retrieve risk summary data from the combined Excel file
        """
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            # Access the "Combined Data" sheet
            if "Combined Data" in wb.sheetnames:
                sheet = wb["Combined Data"]
            else:
                # If not found, use the first sheet
                sheet = wb.active
            
            # Read all rows from the sheet
            all_rows = []
            for row in sheet.iter_rows(values_only=True):
                all_rows.append(row)
            
            # Count risk levels
            before_counts, after_counts = self.count_risk_levels(all_rows)
            
            if show == 'before':
                return before_counts
            else:
                return after_counts

        except Exception as e:
            print(f"Error loading risk summary data: {e}")
            return {'H': 0, 'M': 0, 'L': 0}  # Return default values
    

    def create_pie_chart(self, sheet, title, data, position):
        """
        Crea un gràfic de sectors (pie chart) a la fulla especificada amb format net:
        sense etiquetes, només títol i llegenda amb bon espaiat.
        """
        labels = ['High Risk (H)', 'Medium Risk (M)', 'Low Risk (L)']
        values = [data.get('H', 0), data.get('M', 0), data.get('L', 0)]
        start_row = position[0]
        start_col = position[1]

        # Escriure dades a la fulla
        sheet.cell(row=start_row, column=start_col, value="Risk Level")
        sheet.cell(row=start_row, column=start_col+1, value="Count")
        for i, (label, value) in enumerate(zip(labels, values)):
            sheet.cell(row=start_row+1+i, column=start_col, value=label)
            sheet.cell(row=start_row+1+i, column=start_col+1, value=value)

        # Crear el gràfic
        chart = PieChart()
        chart.style = 10
        chart.width = 12
        chart.height = 10  # Suficient per evitar solapaments

        # Configurar les dades
        labels_ref = Reference(sheet, min_col=start_col, min_row=start_row+1, max_row=start_row+3)
        data_ref = Reference(sheet, min_col=start_col+1, min_row=start_row+1, max_row=start_row+3)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(labels_ref)

        chart.title = title
        chart.legend.position = 'b'  # Bottom

        # Afegir el gràfic a la fulla
        chart_anchor = f"{chr(65 + start_col + 2)}{start_row}"
        sheet.add_chart(chart, chart_anchor)

        return chart
    

    def display_risk_pie_chart(self, data, title="Risk Distribution"):
        """
        Create a clean, professional pie chart with no axes
        and properly spaced legend
        """
        labels = ['High Risk (H)', 'Medium Risk (M)', 'Low Risk (L)']
        values = [data.get('H', 0), data.get('M', 0), data.get('L', 0)]
        colors = ["#C40000", "#FFA600", "#00BB00"]  # Red, Orange, Green
        
        # Create figure without axes
        fig = Figure(figsize=(5, 5), facecolor='none')
        ax = fig.add_subplot(111)
        
        ax.axis('off')      # Remove axes completely
        total = sum(values) # Calculate total
        
        if total > 0:
            # Create pie chart
            wedges, texts, autotexts = ax.pie(
                values, 
                colors=colors,
                autopct=lambda p: f'{p:.1f}%' if p > 0 else '',
                startangle=90,
                textprops={'fontsize': 10, 'color': 'white', 'fontweight': 'bold'},
                wedgeprops={'edgecolor': 'black', 'linewidth': 0.5},
                pctdistance=0.7  # Move percentages closer to edge
            )
            
            # Add title
            ax.set_title(title, fontsize=10, fontweight='bold', pad=5)
            
            # Add legend below the pie chart with proper spacing
            legend = ax.legend(
                wedges,
                labels,
                loc='upper center',
                bbox_to_anchor=(0.5, -0.05),  # Position closer to chart
                frameon=False,
                fontsize=8,
                ncol=3
            )
            
            # Add minimal padding between chart and legend
            fig.tight_layout(pad=1.0)
        else:
            # Display message
            ax.text(0.5, 0.5, "No Risk Data Available", 
                    ha='center', va='center', fontsize=12)
        
        return fig

    
    def combine_amfe_data(self, amfe_paths, output_file):
        """
        Combina dades de múltiples AMFEs i crea un nou fitxer Excel amb gràfics millorats
        """
        # Obtenir tots els fitxers Excel
        amfe_files = self.get_amfe_excels(amfe_paths)
        
        print(f"Found {len(amfe_files)} AMFE files to process")
        
        # Llegir dades de cada AMFE
        all_rows = []   # Llista plana de files
        valid_files = 0
        
        for i, file_path in enumerate(amfe_files):
            print(f"Processing file {i+1}/{len(amfe_files)}: {file_path}")
            
            # Include header only for the first valid file
            include_header = (valid_files == 0)
            file_data = self.read_amfe_data(file_path, include_header)
            
            if file_data:
                valid_files += 1
                # Afegir totes les files d'aquest fitxer
                all_rows.extend(file_data)
                
                # Afegir una fila buida de separació (excepte després de l'últim fitxer)
                if i < len(amfe_files) - 1:
                    num_cols = len(file_data[0])  # Agafem el nombre de columnes de la primera fila
                    all_rows.append([None] * num_cols)
        
        print(f"Successfully processed {valid_files}/{len(amfe_files)} files")
        
        if valid_files == 0:
            print("No valid data found to export")
            return None
        
        # Crear nou workbook
        wb = openpyxl.Workbook()
        
        # Eliminar full per defecte
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Crear els tres fulls requerits
        graphics_sheet = wb.create_sheet("Full Process Graphics")
        wb.create_sheet("Individual Charts")
        combined_sheet = wb.create_sheet("Combined Data")
        
        # Escriure dades combinades al full "Combined Data"
        for row_index, row in enumerate(all_rows, 1):
            for col_index, value in enumerate(row, 1):
                combined_sheet.cell(row=row_index, column=col_index, value=value)
        
        # Comptar nivells de risc
        before_counts, after_counts = self.count_risk_levels(all_rows)
        print(f"Risk counts - Before: {before_counts}, After: {after_counts}")
        
        # Configurar fulla de gràfics
        graphics_sheet.title = "Full Process Graphics"
        
        # Afegir títol a la fulla de gràfics
        title_cell = graphics_sheet['A1']
        title_cell.value = "Risk Distribution Analysis"
        title_cell.font = openpyxl.styles.Font(bold=True, size=16)
        
        # Afegir espai entre títol i gràfics
        graphics_sheet.row_dimensions[3].height = 30
        
        # Crear gràfics amb més espai entre ells
        self.create_pie_chart(
            graphics_sheet, 
            "Initial Risk Distribution (Before Actions)", 
            before_counts, 
            (3, 1)  # Fila 3, Columna A
        )
        
        self.create_pie_chart(
            graphics_sheet, 
            "Residual Risk Distribution (After Actions)", 
            after_counts, 
            (3, 7)  # Fila 3, Columna G (més espai)
        )
        
        # Ajustar amplada de columnes per a la fulla de gràfics
        for col in range(1, 15):
            col_letter = openpyxl.utils.get_column_letter(col)
            graphics_sheet.column_dimensions[col_letter].width = 15
        
        # Guardar el nou fitxer
        wb.save(output_file)
        print(f"Combined file with charts saved to: {output_file}")
        
        # Retornar recomptes per a la UI
        return {
            'output_file': output_file,
            'before_counts': before_counts,
            'after_counts': after_counts
        }

# Exemple d'ús amb gestió d'errors
if __name__ == "__main__":
    amfe_manager = AmfeManager()

    # Simular selecció d'usuari
    selected_processes = ['09_Resistance_Welding+Rivetting (SQB)','01_Deburring_hard (SQB) I']

    # Obtenir camins AMFE
    amfe_paths = amfe_manager.get_amfe_paths(selected_processes)

    # Combinar dades i guardar
    output_path = "Combined_AMFE_Output.xlsx"
    result = amfe_manager.combine_amfe_data(amfe_paths, output_path)

    if result:
        print(f"Fitxer combinat creat amb èxit: {result['output_file']}")
        print(f"Recompte de risc abans: {result['before_counts']}")
        print(f"Recompte de risc després: {result['after_counts']}")
    else:
        print("No s'ha pogut crear el fitxer combinat")
